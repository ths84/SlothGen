from bs4 import BeautifulSoup
import requests
import openpyxl as xl


def web_scrape_first_names(wb_filename):
    wb = xl.load_workbook(wb_filename)
    sheet = wb.active
    url_root = 'https://www.vorname.com'
    url_part2 = ['maedchennamen', 'jungennamen']
    for part_in_url_part2 in url_part2:
        # Get starting row in Excel sheet to append new names to existing ones
        column = 1
        for starting_row in sheet.iter_cols(min_row=1, min_col=1, max_col=1):
            for cell in starting_row:
                if cell.value is None:
                    starting_row = cell.row
                elif cell.value is not None:
                    starting_row = cell.row + 1

        # Loop through all first names starting from 'A' to 'Z'
        for url_part3 in range(ord('A'), ord('Z') + 1):
            wb.save(wb_filename)
            url_combine = f'{url_root}/{part_in_url_part2},{chr(url_part3)},1.html'
            response = requests.get(url_combine)
            if response.status_code == 200:
                html_source = requests.get(url_combine).text
                soup = BeautifulSoup(html_source, 'lxml')
                pagination = soup.find('div', class_='pagination')

                # Get all page-urls
                links = [link['href'] for link in pagination.find_all('a', href=True)]

                checker = None
                while checker is None:
                    html_source2 = requests.get(f'{url_root}/{links[-1]}').text
                    soup = BeautifulSoup(html_source2, 'lxml')
                    pagination = soup.find('div', class_='pagination')
                    checker = soup.find('div', class_='next disabled')
                    if checker is None:
                        for link in pagination.find_all('a', href=True):
                            links.append(link['href'])
                    else:
                        break

                # Get all first_names from all page-urls
                for link in links:
                    html_source = requests.get(f'{url_root}/{link}').text
                    soup = BeautifulSoup(html_source, 'lxml')
                    female_first_names = soup.find_all('td', class_='name')
                    for name in female_first_names:
                        cell = sheet.cell(starting_row, column)
                        cell.value = name.text
                        starting_row += 1
                        print(f'... {starting_row} ... {cell.value}')
                print(f'Letter {chr(url_part3)} ... DONE ... SAVING.')
            else:
                print(f'{url_combine} not on server... continuing...')

    wb.save(wb_filename)