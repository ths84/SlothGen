from bs4 import BeautifulSoup
import requests
import re
import openpyxl as xl


def scraping_last_names_wikipedia(wb_filename):
    url = 'https://de.wiktionary.org/wiki/Verzeichnis:Deutsch/Namen/die_häufigsten_Nachnamen_Deutschlands'
    html_source = requests.get(url).text
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(html_source, 'lxml')
        content_block = soup.find('ol')

        wb = xl.load_workbook(wb_filename)
        sheet = wb["sheet1"]
        row = 1
        column = 3
        for name in content_block.find_all('a', title=True):
            cell = sheet.cell(row, column)
            cell.value = name.text
            row += 1
            print(f'... {row} {cell.value}')
    else:
        print(f'{url} not on server... continuing...')

    wb.save(wb_filename)


def scraping_last_names_familyeducationdotcom(wb_filename):
    wb = xl.load_workbook(wb_filename)
    sheet = wb["sheet1"]
    column = 3
    # Get starting row in Excel sheet to append new names to existing ones
    for row in sheet.iter_cols(min_row=1, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                starting_row = cell.row

    url_root = 'https://www.familyeducation.com/baby-names/browse-names/surname'
    for url_appendix in range(ord('a'), ord('z') + 1):
        wb.save(wb_filename)
        url_combine = f'{url_root}/{chr(url_appendix)}'
        html_source = requests.get(url_combine).text
        response = requests.get(url_combine)
        if response.status_code == 200:
            soup = BeautifulSoup(html_source, 'lxml')

            # Find page links on site for set surname starting with url_appendix
            pagination = soup.find('ul', {'class': 'pager__items js-pager__items pagination'})

            if pagination is not None:
                # Get page links from first site for set surname starting with url_appendix
                links = [link['href'] for link in pagination.find_all('a', href=True)]
                for i in range(1, 3):
                    del links[-1]

                # Get last available page link for set surname starting with url_appendix
                pagination_last = soup.find('li', {'class': 'pager__item pager__item--last hide-li'})
                last_page_url = pagination_last.find('a', href=True)
                links.append(last_page_url['href'])

                # Fill in missing page links
                first_string = links[-2]
                last_string = links[-1]
                num_first_string = int(re.search(r'\d+', first_string).group())
                num_last_string = int(re.search(r'\d+', last_string).group())

                missing_links = []
                for i in range(num_first_string, num_last_string):
                    missing_links.append(f'?page={i}')

                in_links = set(links)
                in_missing_links = set(missing_links)
                in_links_but_not_in_missing_links = in_missing_links - in_links
                complete_links = links + list(in_links_but_not_in_missing_links)

                # Get names from all pages for set surname
                for link in complete_links:
                    html_source = requests.get(f'{url_root}/{chr(url_appendix)}{link}').text
                    soup = BeautifulSoup(html_source, 'lxml')
                    section = soup.find('section', {'id': 'block-fentheme-content'})
                    names_list = section.find_all('a', {'href': re.compile("/baby-names/name-meaning/")})
                    for name in names_list:
                        last_name = name.text
                        cell = sheet.cell(starting_row, column)
                        cell.value = name.text
                        starting_row += 1
                        print(f'... {starting_row} ... {cell.value}')
            else:
                print(f'No pagination found on site; continuing with name scrape ...')
                html_source = requests.get(f'{url_root}/{chr(url_appendix)}').text
                soup = BeautifulSoup(html_source, 'lxml')
                section = soup.find('section', {'id': 'block-fentheme-content'})
                names_list = section.find_all('a', {'href': re.compile("/baby-names/name-meaning/")})
                for name in names_list:
                    last_name = name.text
                    cell = sheet.cell(starting_row, column)
                    cell.value = name.text
                    starting_row += 1
                    print(f'... {starting_row} ... {cell.value}')

        else:
            print(f'{url_root} not on server... continuing...')

    wb.save(wb_filename)